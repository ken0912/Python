from openpyxl import Workbook,load_workbook
import re
import xlwt

workbook = load_workbook('D:\Work\Project\SEO\JobMatchSentence_20190717.xlsx')
# sheets = workbook.get_sheet_names()         #从名称获取sheet
booksheet = workbook.get_sheet_by_name('Sheet1')
#booksheet = workbook.active

rows = booksheet.rows
columns = booksheet.columns
#迭代所有的行
l=[]
for row in rows:
    line = [col.value for col in row]
    l.append(line)

fl=[]
a=0
for i in l:
    if re.search(r'\d{3}\-\d{3}\-\d{4}',i[1]):
        a+=1
    else:
        fl.append(i)
print("find {0} sentence".format(a))
print((fl))

def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")



write_excel_xlsx('result2019-07-18.xlsx', 'sheet1', fl)